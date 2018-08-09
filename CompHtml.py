from bs4 import BeautifulSoup
import requests
import openpyxl as px
import os


def fetchHTML(url, cls, flag=0):
    res = requests.get(url, verify=False)
    soup = BeautifulSoup(res.text, 'lxml')
    part = None
    outputDir = ''

    if soup.find(class_=cls) is not None:
        #print(url, '\n')
        #print(cls, '\n')
        part =  soup.find(class_=cls).prettify()

    if flag == 1:
        url = url.split('/')[3]
        outputDir = 'CompHtmlOutput/honban/{}_{}.txt'.format(url, cls)
    else:
        url = url.split('/')[3]
        outputDir = 'CompHtmlOutput/kensho/{}_{}.txt'.format(url, cls)

    #import pdb; pdb.set_trace()
    if part is not None:
        with open(outputDir, 'w', encoding='UTF-8') as f:
            f.write(part)
            #f.write(re.sub("^[\r\n]+", "", part)) # 引数の文字列をファイルに書き込む

def makeDic(target_dictionary, target_path):
    for root, dirs, files in os.walk(target_path):
        for name in files:
            path = os.path.join(root, name)
            with open(path, encoding = 'utf-8')as f:
                target_dictionary[name] = f.read()

def getExceData(list_, sheet_):
    for cell_obj in list(sheet_.columns)[1]:
        list_.append(cell_obj.value)
    list_.pop(0)
    return list_


def main():
    clsArray = ['l-headerGnav', 'l-headerMain',  'm-navSubmenu', 'l-footer']
    inputFile = 'CompHtmlInput/urlList.xlsx'

    wb = px.load_workbook(inputFile)
    prdsheet = wb['Sheet1']
    testsheet = wb['Sheet2']
    prdList = []
    testList = []

    prdList = getExceData(prdList, prdsheet)
    testList = getExceData(testList, testsheet)

    for url in prdList:
        for cls in clsArray:
            part = fetchHTML(url, cls, 1)

    for url in testList:
        for cls in clsArray:
            part = fetchHTML(url, cls)

    print('HTML内容ファイルの吐き出しが完了しました。')

    honbanDic = {}
    kenshoDic = {}
    #import pdb; pdb.set_trace()
    honban_path = 'CompHtmlOutput/honban'
    kensho_path = 'CompHtmlOutput/kensho'
    makeDic(honbanDic, honban_path)
    makeDic(kenshoDic, kensho_path)

    outputFile = 'CompHtmlOutput/output.xlsx'
    outputWb = px.Workbook()
    outputWs = outputWb['Sheet']
    font = px.styles.Font(name = 'Meiryo UI')

    for row, key in enumerate(honbanDic):
        row += 1
        keyCell = outputWs.cell(row=row, column=1, value=key)
        keyCell.font = font
        if honbanDic.get(key) == kenshoDic.get(key):
            cell = outputWs.cell(row = row, column = 2, value = '○')
            cell.font = font
        else:
            cell = outputWs.cell(row = row, column = 2, value = '×')
            cell.font = font

    outputWb.save(outputFile)
    print('エクセルファイルの吐き出しが完了しました。')

    """
    for col in outputWb.iter_rows(min_row=1, max_col=2, max_row=len(honbanDic)):
        for cell in col:
            cell = key
            for key in honbanDic.keys():
                cell = key
                if honbanDic.get(key) == kenshoDic.get(key):
                    print(key)
                    print('一致です。')
                else:
                    print(key)
                    print('不一致です。')
    outputWb.save(outputFile)
    """


if __name__ == '__main__':
    main()
