import openpyxl as px
import os

path = r'C:\Users\genki.naito\Downloads\8月2日分_最新版 (1)\8月2日分_最新版'
file_list = []
cell_list = []




for root, dirs, files in os.walk(path):
    for file in files:
        if file.startswith('カテゴリー'):
            target_file = os.path.join(root, file)
            print(target_file)
            file_list.append(target_file)

for target_file in file_list:
    print('ループ')
    print(target_file)
    wb = px.load_workbook(target_file , read_only=True)
    ws = wb['master']
    ranges = ws.calculate_dimension()
    print(ranges)
    print('レンジ取得')
    cells = ws[ranges]
    cell_list.append(cells)
    #print(cells)
    break

"""
outputFile = 'CompHtmlOutput/output.xlsx'
outputWb = px.Workbook()
outputWs = outputWb['Sheet']
font = px.styles.Font(name = 'Meiryo UI')

for row, key in enumerate(honbanDic):
    row += 1
    keyCell = outputWs.cell(row=row, column=1, value=key)
    keyCell.font = font
    if honbanDic.get(key) == kenshoDic.get(key):
        cell = outputWs.cell(row = row, column = 2, value = '○')
        cell.font = font
    else:
        cell = outputWs.cell(row = row, column = 2, value = '×')
        cell.font = font

outputWb.save(outputFile)
"""
